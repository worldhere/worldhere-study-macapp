import re
from typing import List, Dict, Optional, Tuple
from openpyxl import load_workbook

# 字段名映射：目标字段 → 可识别的列名变体（大小写不敏感）
FIELD_ALIASES: Dict[str, List[str]] = {
    "name": [
        "任务名", "任务名称", "采集任务名称", "名称", "name", "task name", "task_name",
        "任务", "task",
    ],
    "type": [
        "机型", "机器类型", "设备类型", "type", "machine type", "machine_type",
        "类型", "设备", "machine",
    ],
    "task_kind": [
        "任务类型", "采集类型", "task kind", "task_kind", "kind", "task_type",
        "任务类别",
    ],
    "priority": [
        "优先级", "priority", "pri", "优先", "优先度",
    ],
    "difficulty": [
        "难度", "difficulty", "diff", "困难度",
    ],
    "duration": [
        "预估时长", "预计时长", "duration", "时长", "est", "预估", "预计",
    ],
    "rbp_task_id": [
        "RBP数采任务ID", "RBP 数采任务 ID", "rbp_task_id", "rbp id", "rbp",
        "任务ID", "task id", "task_id", "RBP任务ID",
    ],
    "scene": [
        "场景", "任务场景", "scene", "应用场景",
    ],
    "general_category": [
        "通用类别", "通用任务类别", "general_category", "category", "类别",
        "通用分类",
    ],
    "source_link": [
        "来源链接", "任务来源链接", "source_link", "link", "url", "链接",
        "来源URL", "任务链接",
    ],
    "expected_count": [
        "预期采集量", "预期采集量/条", "预期采集条数", "任务条数", "条数",
        "expected_count", "expcnt", "count", "数量", "采集量", "采集条数",
        "预期数量",
    ],
    "collection_req_id": [
        "数采需求ID", "数采需求 ID", "collection_req_id", "creqid",
        "需求ID", "采集需求ID",
    ],
    "collection_req_type": [
        "数采需求类型", "collection_req_type", "creqtype",
        "需求类型", "采集需求类型",
    ],
    "remark": [
        "备注", "remark", "note", "说明", "备注信息", "描述",
    ],
    "package_name": [
        "所属任务包", "任务包", "package", "package_name", "所属包", "包名",
        "任务包名称",
    ],
    "package_deadline": [
        "截止时间", "截止日期", "deadline", "due", "到期",
    ],
}


def _normalize(s: str) -> str:
    """去除空格、下划线、大小写差异，用于模糊比较"""
    return re.sub(r"[\s_]+", "", str(s).lower())


def detect_fields(headers: List[str]) -> Dict[str, int]:
    """
    根据表头行自动识别字段映射。
    返回 {目标字段名: 列索引(0-based)}
    """
    mapping: Dict[str, int] = {}
    used_aliases: Dict[str, str] = {}  # target_field -> matched_header_text

    for idx, raw_header in enumerate(headers):
        if raw_header is None:
            continue
        h = str(raw_header).strip()
        if not h:
            continue
        norm_h = _normalize(h)

        best_field = None
        best_score = 0

        for field, aliases in FIELD_ALIASES.items():
            if field in mapping:
                continue  # 已经匹配到了
            for alias in aliases:
                norm_a = _normalize(alias)
                # 完全匹配
                if norm_h == norm_a:
                    score = 100
                # 包含匹配
                elif norm_a in norm_h or norm_h in norm_a:
                    score = 80
                else:
                    continue
                if score > best_score:
                    best_score = score
                    best_field = field

        if best_field and best_score >= 80:
            mapping[best_field] = idx
            used_aliases[best_field] = h

    return mapping


def _safe_str(val) -> str:
    if val is None:
        return ""
    if isinstance(val, float):
        if val == int(val):
            return str(int(val))
    if isinstance(val, int):
        return str(val)
    return str(val).strip()


def _safe_int(val) -> Optional[int]:
    if val is None:
        return None
    try:
        return int(float(str(val).strip()))
    except (ValueError, TypeError):
        return None


def _cell_to_str(val) -> str:
    """将 Excel 单元格值转为规范字符串。数字去掉无意义的 .0 后缀。"""
    if val is None:
        return ""
    if isinstance(val, float):
        if val == int(val):
            return str(int(val))
        return str(val)
    if isinstance(val, int):
        return str(val)
    return str(val).strip()


def list_excel_sheets(file_path: str) -> List[str]:
    """返回 Excel 文件中所有工作表名称"""
    import os
    ext = os.path.splitext(file_path)[1].lower()
    if ext == '.xls':
        import xlrd
        wb = xlrd.open_workbook(file_path)
        return [wb.sheet_by_index(i).name for i in range(wb.nsheets)]
    else:
        wb = load_workbook(file_path, data_only=True, read_only=True)
        names = wb.sheetnames
        wb.close()
        return names


def parse_excel(file_path: str, sheet_name: Optional[str] = None) -> Tuple[Dict[str, int], List[Dict], List[str]]:
    """
    解析 Excel 文件（支持 .xlsx 和 .xls）。
    sheet_name: 指定工作表名，None 表示使用第一个工作表。
    返回: (字段映射, 行数据列表, 原始表头列表)
    """
    import os
    ext = os.path.splitext(file_path)[1].lower()

    if ext == '.xls':
        return _parse_xls(file_path, sheet_name)
    else:
        return _parse_xlsx(file_path, sheet_name)


def _parse_xlsx(file_path: str, sheet_name: Optional[str] = None) -> Tuple[Dict[str, int], List[Dict], List[str]]:
    wb = load_workbook(file_path, data_only=True)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    rows_iter = ws.iter_rows(min_row=1, values_only=True)

    try:
        headers = [_cell_to_str(c) for c in next(rows_iter)]
    except StopIteration:
        return {}, [], []

    if all(h == "" or h == "None" for h in headers):
        try:
            headers = [_cell_to_str(c) for c in next(rows_iter)]
        except StopIteration:
            return {}, [], []

    field_map = detect_fields(headers)

    rows = []
    for row in rows_iter:
        row_vals = [_cell_to_str(c) for c in row]
        if all(v == "" for v in row_vals):
            continue
        item = {}
        for field, idx in field_map.items():
            val = row[idx] if idx < len(row) else None
            item[field] = val
        rows.append(item)

    return field_map, rows, headers


def _parse_xls(file_path: str, sheet_name: Optional[str] = None) -> Tuple[Dict[str, int], List[Dict], List[str]]:
    import xlrd
    wb = xlrd.open_workbook(file_path)
    if sheet_name:
        try:
            ws = wb.sheet_by_name(sheet_name)
        except xlrd.XLRDError:
            ws = wb.sheet_by_index(0)
    else:
        ws = wb.sheet_by_index(0)

    if ws.nrows < 1:
        return {}, [], []

    headers = [_cell_to_str(ws.cell_value(0, c)) for c in range(ws.ncols)]

    # 跳过完全空白的表头行
    if all(h == "" or h == "None" for h in headers):
        if ws.nrows < 2:
            return {}, [], []
        headers = [_cell_to_str(ws.cell_value(1, c)) for c in range(ws.ncols)]

    field_map = detect_fields(headers)

    start_row = 1
    if all(h == "" or h == "None" for h in [_cell_to_str(ws.cell_value(0, c)) for c in range(ws.ncols)]):
        start_row = 2

    rows = []
    for r in range(start_row, ws.nrows):
        row_vals = [_cell_to_str(ws.cell_value(r, c)) for c in range(ws.ncols)]
        if all(v == "" for v in row_vals):
            continue
        item = {}
        for field, idx in field_map.items():
            if idx < ws.ncols:
                item[field] = ws.cell_value(r, idx)
            else:
                item[field] = None
        rows.append(item)

    return field_map, rows, headers


def analyze_import(rows: List[Dict], field_map: Dict[str, int], default_type: str = "BR2") -> Dict:
    """
    分析导入数据：去重检查、字段识别情况。
    返回分析结果供前端展示和确认。
    """
    from db import get_db, get_allowed_task_kinds, get_allowed_machine_types, get_allowed_priorities, get_allowed_difficulties

    conn = get_db()
    # 加载现有任务的索引
    existing_tasks = conn.execute(
        "SELECT id, name, type, rbp_task_id FROM tasks"
    ).fetchall()
    conn.close()

    existing_rbp_ids = set()
    existing_name_type_pairs = set()
    for t in existing_tasks:
        rbp = (t["rbp_task_id"] or "").strip()
        if rbp:
            existing_rbp_ids.add(rbp.lower())
        nt = (_normalize(t["name"]), _normalize(t["type"]))
        existing_name_type_pairs.add(nt)

    items = []
    rbp_dup_count = 0
    name_type_dup_count = 0
    ok_count = 0

    raw_types = set()
    raw_kinds = set()
    raw_priorities = set()
    raw_difficulties = set()

    allowed_kinds = get_allowed_task_kinds()
    allowed_types = get_allowed_machine_types()
    allowed_priorities = get_allowed_priorities()
    allowed_difficulties = get_allowed_difficulties()

    for idx, row in enumerate(rows):
        name = _safe_str(row.get("name"))
        rbp_id = _safe_str(row.get("rbp_task_id"))

        # 任务名和RBP任务ID至少需要一个
        if not name and not rbp_id:
            continue

        status = "ok"
        warnings = []

        # 规则0：任务名为空但RBP任务ID存在 → 待确认
        if not name and rbp_id:
            status = "confirm"
            warnings.append("任务名为空（仅靠RBP任务ID标识），请确认是否导入")

        # 规则1：RBP 任务ID 重复 → 拒绝
        if rbp_id and rbp_id.lower() in existing_rbp_ids:
            status = "rejected"
            warnings.append(f"RBP任务ID「{rbp_id}」已存在，跳过")
            rbp_dup_count += 1

        # 规则2：任务名 + 机型重复 → 待确认
        row_type = _safe_str(row.get("type"))
        if status != "rejected" and row_type:
            nt = (_normalize(name), _normalize(row_type))
            if nt in existing_name_type_pairs:
                status = "confirm"
                warnings.append(f"任务名「{name}」+ 机型「{row_type}」与已有任务重复，需确认是否导入")
                name_type_dup_count += 1

        if status == "ok":
            ok_count += 1

        # 收集原始枚举值
        raw_type = _safe_str(row.get("type"))
        raw_kind = (_safe_str(row.get("task_kind")) or "").strip()
        raw_pri = (_safe_str(row.get("priority")) or "").strip()
        raw_diff = (_safe_str(row.get("difficulty")) or "").strip()
        if raw_type:
            raw_types.add(raw_type)
        if raw_kind:
            raw_kinds.add(raw_kind)
        if raw_pri:
            raw_priorities.add(raw_pri)
        if raw_diff:
            raw_difficulties.add(raw_diff)

        # 规范化数据
        task_kind = raw_kind
        if not task_kind or task_kind not in allowed_kinds:
            task_kind = allowed_kinds[0] if allowed_kinds else "常规"

        row_type = raw_type
        if not row_type or row_type not in allowed_types:
            if row_type:
                t_upper = _normalize(row_type)
                sorted_types = sorted(allowed_types, key=len, reverse=True)
                for at in sorted_types:
                    if _normalize(at) in t_upper:
                        row_type = at
                        break
            if not row_type or row_type not in allowed_types:
                row_type = default_type

        items.append({
            "index": idx,
            "name": name,
            "type": row_type,
            "task_kind": task_kind,
            "priority": _safe_str(row.get("priority")),
            "difficulty": _safe_str(row.get("difficulty")),
            "duration": _safe_str(row.get("duration")),
            "rbp_task_id": rbp_id,
            "scene": _safe_str(row.get("scene")),
            "general_category": _safe_str(row.get("general_category")),
            "source_link": _safe_str(row.get("source_link")),
            "expected_count": _safe_int(row.get("expected_count")),
            "collection_req_id": _safe_str(row.get("collection_req_id")),
            "collection_req_type": _safe_str(row.get("collection_req_type")),
            "remark": _safe_str(row.get("remark")),
            "package_name": _safe_str(row.get("package_name")),
            "status": status,
            "warnings": warnings,
        })

    # 字段映射描述（给前端看）
    field_labels = []
    for field, idx in field_map.items():
        field_labels.append({
            "field": field,
            "col_index": idx,
            "matched_header": str(list(FIELD_ALIASES.keys())[list(FIELD_ALIASES.keys()).index(field)] if field in FIELD_ALIASES else field),
        })

    # 检测导入数据中存在但设置中不存在的枚举值
    missing_types = {}
    if raw_types - set(allowed_types):
        missing_types["type"] = sorted(raw_types - set(allowed_types))
    if raw_kinds - set(allowed_kinds):
        missing_types["task_kind"] = sorted(raw_kinds - set(allowed_kinds))
    if raw_priorities - set(allowed_priorities):
        missing_types["priority"] = sorted(raw_priorities - set(allowed_priorities))
    if raw_difficulties - set(allowed_difficulties):
        missing_types["difficulty"] = sorted(raw_difficulties - set(allowed_difficulties))

    return {
        "total_rows": len(rows),
        "valid_items": len(items),
        "ok_count": ok_count,
        "rbp_dup_count": rbp_dup_count,
        "name_type_dup_count": name_type_dup_count,
        "field_map": {f: i for f, i in field_map.items()},
        "items": items,
        "missing_types": missing_types,
    }


def execute_import(items_to_import: List[Dict], package_name: Optional[str] = None,
                   package_deadline: Optional[str] = None, machine_type: str = "BR2") -> Dict:
    """
    执行导入。items_to_import 为前端确认后的任务列表。
    后端再次校验 RBP ID 去重。
    当 package_name 提供时，自动创建任务包并将任务关联。
    """
    from db import get_db, get_allowed_task_kinds, get_allowed_machine_types, get_allowed_machine_types
    from utils import parse_duration_to_minutes

    conn = get_db()

    package_id = None
    if package_name:
        import datetime as _dt
        pcur = conn.execute(
            "INSERT INTO task_packages(name, deadline, priority, machine_type, created_at) VALUES (?,?,?,?,?)",
            (package_name, package_deadline, "P1", machine_type, _dt.datetime.now().isoformat(timespec="seconds")),
        )
        package_id = pcur.lastrowid
    # 加载现有 RBP ID 用于后端二次校验
    existing = conn.execute("SELECT rbp_task_id FROM tasks WHERE rbp_task_id != ''").fetchall()
    existing_rbp_ids = set((t["rbp_task_id"] or "").strip().lower() for t in existing)

    imported = 0
    skipped = 0
    errors = []

    for item in items_to_import:
        name = _safe_str(item.get("name"))
        rbp_id = _safe_str(item.get("rbp_task_id"))
        if not name and not rbp_id:
            skipped += 1
            continue
        if not name and rbp_id:
            name = "[RBP:" + rbp_id[:20] + "]"

        rbp_id = _safe_str(item.get("rbp_task_id"))
        if rbp_id and rbp_id.lower() in existing_rbp_ids:
            skipped += 1
            errors.append(f"「{name}」的RBP任务ID「{rbp_id}」已存在，跳过")
            continue

        allowed_types = get_allowed_machine_types()
        row_type = _safe_str(item.get("type"))
        if not row_type or row_type not in allowed_types:
            if machine_type and machine_type in allowed_types:
                row_type = machine_type
            else:
                row_type = allowed_types[0] if allowed_types else "BR1"

        allowed_kinds = get_allowed_task_kinds()
        task_kind = (_safe_str(item.get("task_kind")) or "").strip()
        if not task_kind or task_kind not in allowed_kinds:
            task_kind = allowed_kinds[0] if allowed_kinds else "常规"

        duration_str = _safe_str(item.get("duration"))
        est_mode = "blank"
        est_seconds = None
        if duration_str:
            m = parse_duration_to_minutes(duration_str, default_minutes=0)
            if m:
                est_mode = "direct"
                est_seconds = m * 60
        elif name and not duration_str:
            from utils import estimate_duration_from_name
            count = _safe_int(item.get("expected_count")) or 25
            est_min = estimate_duration_from_name(name, count)
            duration_str = str(est_min)
            est_mode = "auto"
            est_seconds = est_min * 60

        difficulty_str = _safe_str(item.get("difficulty"))
        if not difficulty_str and name:
            from utils import estimate_difficulty_from_name
            difficulty_str = estimate_difficulty_from_name(name)

        try:
            conn.execute(
                "INSERT INTO tasks(name,type,task_kind,priority,difficulty,duration,est_mode,est_seconds,"
                "remark,status,rbp_task_id,scene,general_category,source_link,"
                "expected_count,collection_req_id,collection_req_type,package_id) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (
                    name,
                    row_type,
                    task_kind,
                    _safe_str(item.get("priority")) or ("P1" if package_name else ""),
                    difficulty_str,
                    duration_str,
                    est_mode,
                    est_seconds,
                    _safe_str(item.get("remark")),
                    "待分配",
                    rbp_id,
                    _safe_str(item.get("scene")),
                    _safe_str(item.get("general_category")),
                    _safe_str(item.get("source_link")),
                    _safe_int(item.get("expected_count")),
                    _safe_str(item.get("collection_req_id")),
                    _safe_str(item.get("collection_req_type")),
                    package_id,
                ),
            )
            imported += 1
            if rbp_id:
                existing_rbp_ids.add(rbp_id.lower())
        except Exception as e:
            errors.append(f"导入「{name}」失败: {e}")
            skipped += 1

    conn.commit()
    conn.close()
    return {"imported": imported, "skipped": skipped, "errors": errors}
